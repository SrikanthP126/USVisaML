import os
import json
import uuid
import calendar
from datetime import datetime
from openpyxl import load_workbook
import argparse
import sys
import shlex
import logging
import hashlib
from pathlib import Path

class ExcelProcessor:
    # Fixed paths relative to script location
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
    LOG_DIR = os.path.join(SCRIPT_DIR, '.logs')
    LOG_FILE = os.path.join(LOG_DIR, 'conversion_history.log')
    HISTORY_FILE = os.path.join(LOG_DIR, '.conversion_registry')

    def __init__(self, input_file, output_directory, force=False):
        self.input_file = os.path.abspath(input_file)
        self.output_directory = os.path.abspath(output_directory)
        self.force = force
        self._ensure_log_directory()
        self.setup_logging()
        
    def _ensure_log_directory(self):
        """Ensure log directory exists and is hidden"""
        os.makedirs(self.LOG_DIR, exist_ok=True)
        # Make directory hidden on Unix systems
        if os.name == 'posix':
            try:
                import subprocess
                subprocess.run(['chattr', '+i', self.LOG_DIR], check=False)
            except:
                pass

    def setup_logging(self):
        formatter = logging.Formatter(
            '%(asctime)s - %(levelname)s - [%(input_file)s] %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        
        # Setup file handler
        self.file_handler = logging.FileHandler(self.LOG_FILE, mode='a')
        self.file_handler.setFormatter(formatter)
        
        # Setup console handler
        self.console_handler = logging.StreamHandler()
        self.console_handler.setFormatter(formatter)
        
        # Setup logger with custom adapter for input file context
        logger = logging.getLogger('ExcelConverter')
        logger.setLevel(logging.INFO)
        logger.addHandler(self.file_handler)
        logger.addHandler(self.console_handler)
        
        # Create adapter to include input file in all log messages
        self.logger = logging.LoggerAdapter(logger, {'input_file': os.path.basename(self.input_file)})

    def get_file_hash(self):
        """Calculate SHA-256 hash of input file"""
        hasher = hashlib.sha256()
        with open(self.input_file, 'rb') as f:
            for chunk in iter(lambda: f.read(4096), b''):
                hasher.update(chunk)
        return hasher.hexdigest()

    def check_previous_conversion(self):
        """Check if file was previously converted"""
        current_hash = self.get_file_hash()
        
        if os.path.exists(self.HISTORY_FILE):
            with open(self.HISTORY_FILE, 'r') as f:
                for line in f:
                    try:
                        file_path, hash_value, timestamp = line.strip().split(',')
                        if file_path == self.input_file and hash_value == current_hash:
                            if not self.force:
                                self.logger.warning(
                                    f"File was already converted on {timestamp}. "
                                    "Use --force to convert again."
                                )
                                return True
                    except ValueError:
                        continue
        return False

    def record_conversion(self):
        """Record successful conversion with timestamp"""
        current_hash = self.get_file_hash()
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        with open(self.HISTORY_FILE, 'a') as f:
            f.write(f"{self.input_file},{current_hash},{timestamp}\n")

    def format_iso_date(self, excel_date):
        try:
            if isinstance(excel_date, str):
                if len(excel_date) == 7:
                    year, month = excel_date.split('-')
                    last_day = calendar.monthrange(int(year), int(month))[1]
                    excel_date = f"{year}-{month}-{last_day}"
                return datetime.strptime(excel_date, "%Y-%m-%d").isoformat() + "-05:00"
        except (ValueError, TypeError):
            return None

    def get_file_tag(self, file_name):
        ext = os.path.splitext(file_name)[1]
        if not ext:
            ext = file_name.split(".")[-1]
        else:
            ext = ext[1:]
        return ext

    def write_records_to_file(self, output_file_path, records, sheet_name):
        upload_count = sum(1 for record in records if record['operation'] == 'upload_new_file')
        with open(output_file_path, 'w') as json_file:
            for record in records:
                json_file.write(json.dumps(record, separators=(',', ':')) + '\n')
        
        self.logger.info(f"Created manifest file: {os.path.basename(output_file_path)}")
        self.logger.info(f"  → Contains {upload_count} upload_new_file operations")
        
        # Log upload files for this sheet
        upload_files = [
            record['file_metadata']['dz_file_name']
            for record in records
            if record['operation'] == 'upload_new_file'
        ]
        self.logger.info(f"Upload files created for sheet {sheet_name}:")
        for file in upload_files:
            self.logger.info(f"  → {file}")
        
        return upload_count

    def process_excel_file(self):
        self.logger.info(f"\n=== Starting to process Excel file: {self.input_file} ===\n")
        
        if self.check_previous_conversion():
            return
            
        try:
            wb = load_workbook(self.input_file, data_only=True)
        except Exception as e:
            self.logger.error(f"Error loading workbook: {e}")
            sys.exit(1)
        
        excel_base_name = os.path.splitext(os.path.basename(self.input_file))[0]
        file_count = 0
        total_upload_operations = 0
        
        metadata_sheets = [sheet for sheet in wb.sheetnames if sheet.startswith("Metadata")]
        if not metadata_sheets:
            self.logger.error("No metadata sheets found in the workbook!")
            sys.exit(1)

        self.logger.info(f"Found {len(metadata_sheets)} metadata sheet(s) to process\n")

        for sheet_name in metadata_sheets:
            self.logger.info(f"\nProcessing sheet: {sheet_name}")
            sheet = wb[sheet_name]
            current_records = []
            records_per_file = 100
            
            publisher = sheet['B1'].value
            region = sheet['R4'].value
            record_class = sheet['C9'].value
            provenance = sheet['D1'].value
            security_classification = sheet['D4'].value
            
            security_classification_map = {
                "Confidential": "C",
                "Highly Confidential": "HC",
                "Internal": "I",
                "Public": "P"
            }
            
            security_classification = security_classification_map.get(security_classification, "I")
            
            creator = sheet['B2'].value
            contributor = sheet['D2'].value
            cost_center = str(sheet['B3'].value).zfill(12) if isinstance(sheet['B3'].value, (int, float)) else sheet['B3'].value.strip()

            row_count = 0
            for row in sheet.iter_rows(min_row=9, values_only=True):
                if row[0] is None:
                    break
                    
                row_count += 1
                if row_count % 10 == 0:
                    self.logger.info(f"  Processing row {row_count}...")
                    
                file_name = row[0]
                file_path = row[1]
                folder_path = f"/dropzone/a360root/{publisher.lower()}/submission/{file_path}/"
                record_code = row[2]
                record_date = row[3]
                formatted_record_date = self.format_iso_date(record_date)
                relation_id = str(uuid.uuid4())
                submission_date = datetime.now().isoformat() + "-05:00"
                
                record_metadata = {
                    "record_class": record_class,
                    "publisher": publisher,
                    "region": region,
                    "recorddate": formatted_record_date,
                    "provenance": provenance,
                    "submission_date": submission_date,
                    "security_classification": security_classification,
                    "contributor": contributor,
                    "creator": creator,
                    "description": row[4],
                    "title": file_name,
                    "language": "eng",
                    "cost_center": cost_center,
                    "date_range": row[5],
                    "major_description": row[6],
                    "minor_description": row[7],
                    "reference_1": row[8]
                }
                
                file_metadata = {
                    "publisher": publisher,
                    "source_folder_path": file_path,
                    "source_file_name": file_name,
                    "dz_folder_path": folder_path,
                    "dz_file_name": file_name,
                    "file_tag": self.get_file_tag(file_name)
                }
                
                current_records.append({
                    "operation": "create_record",
                    "relation_id": relation_id,
                    "record_metadata": record_metadata
                })
                
                current_records.append({
                    "operation": "upload_new_file",
                    "relation_id": relation_id,
                    "file_metadata": file_metadata
                })
                
                if len(current_records) >= records_per_file * 2:
                    output_file_path = os.path.join(
                        self.output_directory, 
                        f"{excel_base_name}_{sheet_name}_{file_count}.a360"
                    )
                    upload_count = self.write_records_to_file(output_file_path, current_records, sheet_name)
                    total_upload_operations += upload_count
                    current_records = []
                    file_count += 1
                    
            if current_records:
                output_file_path = os.path.join(
                    self.output_directory,
                    f"{excel_base_name}_{sheet_name}_{file_count}.a360"
                )
                upload_count = self.write_records_to_file(output_file_path, current_records, sheet_name)
                total_upload_operations += upload_count
                file_count += 1
                
        self.logger.info(f"\n=== Processing Complete ===")
        self.logger.info(f"Total manifest files created: {file_count}")
        self.logger.info(f"Total upload_new_file operations: {total_upload_operations}")
        
        # Record successful conversion
        self.record_conversion()

def main():
    parser = argparse.ArgumentParser(description="Convert Excel file to JSON format.")
    parser.add_argument("input_file", type=str, help="Path to the Excel file")
    parser.add_argument("output_directory", type=str, help="Directory to save JSON files")
    parser.add_argument("--force", action="store_true", 
                       help="Force conversion even if file was previously converted")
    
    try:
        if len(sys.argv) > 1:
            args = parser.parse_args([arg.strip('"') for arg in sys.argv[1:]])
        else:
            parser.print_help()
            sys.exit(1)
            
        # Create output directory if it doesn't exist
        os.makedirs(args.output_directory, exist_ok=True)
        
        processor = ExcelProcessor(
            args.input_file,
            args.output_directory,
            args.force
        )
        processor.process_excel_file()
        
    except Exception as e:
        print(f"Error: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    main()
