import pandas as pd
import json
import uuid
from openpyxl import load_workbook

# Load the Excel file from a specified path
excel_file_path = "path_to_your_excel_file.xlsx"  # Replace with the actual path to your Excel file
wb = load_workbook(excel_file_path, data_only=True)

# Create an empty list to store JSON output
json_output = []

# Constant metadata fields (fixed values)
record_metadata_template = {
    "record_class": "ADM180",
    "publisher": "IDF",
    "region": "USA",
    "provenance": "IBM Data Factory",
    "entitlement_tag": "Internal",
    "contributor": "PL56162",
    "creator": "PT34732",
    "language": "eng",
}

file_metadata_template = {
    "publisher": "IDF",
    "source_folder_path": "QAC",
    "file_tag": "zip"
}

# Helper function to create a unique relation_id
def generate_relation_id():
    return str(uuid.uuid4())

# Function to extract dynamic data from Excel
def extract_excel_data(sheet):
    dynamic_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            data = {
                "file_path": row[0],        # Column A
                "record_code": row[1],      # Column B
                "cost_center": row[3],      # Column D
                "date_range": row[6],       # Column G
                "minor_description": row[7],# Column H
                "major_description": row[8],# Column I
                "record_date": row[9],      # Column K
            }
            dynamic_data.append(data)
    return dynamic_data

# Read all the sheets from the Excel file
sheets = wb.sheetnames

# Process each sheet and generate JSON output
for sheet_name in sheets:
    sheet = wb[sheet_name]
    data_rows = extract_excel_data(sheet)

    for data in data_rows:
        # Generate unique relation ID for each record-operation pair
        relation_id = generate_relation_id()

        # Prepare record_metadata with dynamic data
        record_metadata = record_metadata_template.copy()
        record_metadata.update({
            "recordDate": data["record_date"],
            "title": data["file_path"],
            "Date Range": data["date_range"],
            "Major Description": data["major_description"],
            "Minor Description": data["minor_description"]
        })

        # Prepare file_metadata with dynamic data
        file_metadata = file_metadata_template.copy()
        file_metadata.update({
            "source_file_name": data["file_path"],
            "dz_file_name": data["file_path"]
        })

        # Create the "create_record" entry
        json_output.append({
            "operation": "create_record",
            "relation_id": relation_id,
            "record_metadata": record_metadata
        })

        # Create the "upload_new_file" entry
        json_output.append({
            "operation": "upload_new_file",
            "relation_id": relation_id,
            "file_metadata": file_metadata
        })

# Save the generated JSON output to a file
output_file_path = "path_to_save_your_output.json"  # Replace with your desired output path
with open(output_file_path, 'w') as json_file:
    json.dump(json_output, json_file, indent=4)

print(f"JSON file created successfully at {output_file_path}")