import os
import json
import uuid
import calendar
from openpyxl import load_workbook
from datetime import datetime

# Directory containing multiple Excel files
input_directory = '/ark/landing_zone/exceltosjon/metadata_excel/'
output_directory = '/ark/landing_zone/exceltosjon/metadata_json/'

if not os.path.exists(output_directory):
    os.makedirs(output_directory)

# Helper function to format the date to the last day of the month (if month/year is provided)
def format_iso_date(excel_date):
    try:
        if isinstance(excel_date, str):
            if len(excel_date) == 7:  # Year-Month format (e.g., 2022-02)
                year, month = excel_date.split('-')
                last_day = calendar.monthrange(int(year), int(month))[1]  # Get last day of the month
                excel_date = f"{year}-{month}-{last_day}"  # Construct the full date
                return datetime.strptime(excel_date, "%Y-%m-%d").isoformat() + "-05:00"
    except (ValueError, TypeError):
        return None

# Helper function to get file tag based on file extension
def get_file_tag(file_name):
    ext = os.path.splitext(file_name)[1].lower()
    return {
        '.doc': 'word', '.pdf': 'pdf', '.zip': 'zip', '.jpg': 'image', '.jpeg': 'image',
        '.png': 'image', '.csv': 'csv', '.txt': 'text', '.xlsx': 'excel'
    }.get(ext, 'unknown')

# Helper function to write records to JSON file
def write_records_to_file(output_file_path, records):
    with open(output_file_path, 'w') as json_file:
        for record in records:
            json_file.write(json.dumps(record, separators=(',', ':')) + '\n')

# Function to process each Excel file in the directory
def process_excel_file(file_path):
    print(f"Processing file: {file_path}")
    wb = load_workbook(file_path, data_only=True)
    file_count = 0  # Track the number of output files created for this Excel file

    # Filter only sheets that start with "Metadata"
    metadata_sheets = [sheet_name for sheet_name in wb.sheetnames if sheet_name.startswith("Metadata")]

    # Extract the base name of the Excel file for naming output files
    excel_base_name = os.path.splitext(os.path.basename(file_path))[0]

    for sheet_name in metadata_sheets:
        print(f"Processing sheet: {sheet_name}")
        sheet = wb[sheet_name]
        
        # Read metadata from fixed cells
        publisher = sheet['B1'].value
        region = sheet['B4'].value
        record_class = sheet['C9'].value
        provenance = sheet['D1'].value
        security_classification = sheet['D4'].value

        # Map security classification
        security_classification_map = {
            "Confidential": "C",
            "Highly Confidential": "HC",
            "Internal": "I",
            "Public": "P"
        }
        security_classification = security_classification_map.get(security_classification, "I")

        # Other metadata fields
        creator = sheet['B2'].value
        contributor = sheet['D2'].value
        cost_center = sheet['B3'].value
        cost_center = str(cost_center).zfill(12) if isinstance(cost_center, (int, float)) else cost_center.strip()
        
        file_counter = 1
        current_records = []
        records_per_file = 100

        # Iterate through the rows starting from row 9
        for row in sheet.iter_rows(min_row=9, values_only=True):
            if row[0] is None:
                break  # Stop when an empty row is encountered

            # Extract metadata from the row
            file_name = row[0]
            file_path = row[1]
            record_code = row[2]
            record_date = row[3]
            formatted_record_date = format_iso_date(record_date)
            relation_id = str(uuid.uuid4())
            submission_date = datetime.now().isoformat() + "-05:00"

            # Create the JSON records
            record_metadata = {
                "record_class": record_class,
                "publisher": publisher,
                "region": region,
                "recordDate": formatted_record_date,
                "provenance": provenance,
                "submission_date": submission_date,
                "security_classification": security_classification,
                "contributor": contributor,
                "creator": creator,
                "description": row[4],
                "title": file_name,
                "Language": "eng",
                "cost_center": cost_center,
                "date_range": row[5],
                "major_description": row[6],
                "minor_description": row[7],
                "reference_1": row[8]
            }

            file_metadata = {
                "publisher": publisher,
                "source_folder_path": file_path,
                "source_file_name": file_name,
                "dz_file_name": file_name,
                "file_tag": get_file_tag(file_name)
            }

            # Append records to the list
            current_records.append({
                "operation": "create_record",
                "relation_id": relation_id,
                "record_metadata": record_metadata
            })
            current_records.append({
                "operation": "upload_new_file",
                "relation_id": relation_id,
                "file_metadata": file_metadata
            })

            # Write records to file if enough records are collected
            if len(current_records) >= records_per_file * 2:
                output_file_path = os.path.join(
                    output_directory, f'{excel_base_name}_{sheet_name}_{file_counter}_test.a360'
                )
                write_records_to_file(output_file_path, current_records)
                current_records = []  # Reset for next batch
                file_counter += 1
                file_count += 1

        # Write any remaining records to a final file
        if current_records:
            output_file_path = os.path.join(
                output_directory, f'{excel_base_name}_{sheet_name}_{file_counter}_test.a360'
            )
            write_records_to_file(output_file_path, current_records)
            file_count += 1  # Count the final file

    # Print the total count of files generated for this Excel file
    print(f"Files created for {os.path.basename(file_path)}: {file_count}")

# Function to process files in a directory
def process_directory(directory_path):
    excel_files = [f for f in os.listdir(directory_path) if f.endswith('.xlsx')]
    if not excel_files:
        print("No Excel files found in the input directory.")
    for excel_file in excel_files:
        file_path = os.path.join(directory_path, excel_file)
        process_excel_file(file_path)

    print("All Excel files have been processed. Script will now exit.")

# Main function to trigger the processing
if __name__ == "__main__":
    process_directory(input_directory)
