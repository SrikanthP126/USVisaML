import os
import json
import uuid
import calendar
from datetime import datetime
from openpyxl import load_workbook
import argparse
import sys
import shlex

def format_iso_date(excel_date):
    try:
        if isinstance(excel_date, str):
            if len(excel_date) == 7:
                year, month = excel_date.split('-')
                last_day = calendar.monthrange(int(year), int(month))[1]
                excel_date = f"{year}-{month}-{last_day}"
            return datetime.strptime(excel_date, "%Y-%m-%d").isoformat() + "-05:00"
    except (ValueError, TypeError):
        return None

def get_file_tag(file_name):
    ext = os.path.splitext(file_name)[1]
    if not ext:
        ext = file_name.split(".")[-1]
    else:
        ext = ext[1:]
    return ext

def write_records_to_file(output_file_path, records):
    upload_count = sum(1 for record in records if record['operation'] == 'upload_new_file')
    with open(output_file_path, 'w') as json_file:
        for record in records:
            json_file.write(json.dumps(record, separators=(',', ':')) + '\n')
    print(f"Created manifest file: {os.path.basename(output_file_path)}")
    print(f"  → Contains {upload_count} upload_new_file operations")
    return upload_count

def process_excel_file(file_path, output_directory):
    print(f"\n=== Starting to process Excel file: {file_path} ===\n")
    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        sys.exit(1)
    
    excel_base_name = os.path.splitext(os.path.basename(file_path))[0]
    file_count = 0
    total_upload_operations = 0
    records = []
    current_records = []
    records_per_file = 100

    metadata_sheets = [sheet for sheet in wb.sheetnames if sheet.startswith("Metadata")]
    if not metadata_sheets:
        print("No metadata sheets found in the workbook!")
        sys.exit(1)

    print(f"Found {len(metadata_sheets)} metadata sheet(s) to process\n")

    for sheet_name in metadata_sheets:
        print(f"\nProcessing sheet: {sheet_name}")
        sheet = wb[sheet_name]
        
        publisher = sheet['B1'].value
        region = sheet['R4'].value
        record_class = sheet['C9'].value
        provenance = sheet['D1'].value
        security_classification = sheet['D4'].value
        
        security_classification_map = {
            "Confidential": "C",
            "Highly Confidential": "HC",
            "Internal": "I",
            "Public": "P"
        }
        
        security_classification = security_classification_map.get(security_classification, "I")
        
        creator = sheet['B2'].value
        contributor = sheet['D2'].value
        cost_center = str(sheet['B3'].value).zfill(12) if isinstance(sheet['B3'].value, (int, float)) else sheet['B3'].value.strip()

        row_count = 0
        for row in sheet.iter_rows(min_row=9, values_only=True):
            if row[0] is None:
                break
                
            row_count += 1
            if row_count % 10 == 0:
                print(f"  Processing row {row_count}...")
                
            file_name = row[0]
            file_path = row[1]
            folder_path = f"/dropzone/a360root/{publisher.lower()}/submission/{file_path}/"
            record_code = row[2]
            record_date = row[3]
            formatted_record_date = format_iso_date(record_date)
            relation_id = str(uuid.uuid4())
            submission_date = datetime.now().isoformat() + "-05:00"
            
            record_metadata = {
                "record_class": record_class,
                "publisher": publisher,
                "region": region,
                "recorddate": formatted_record_date,
                "provenance": provenance,
                "submission_date": submission_date,
                "security_classification": security_classification,
                "contributor": contributor,
                "creator": creator,
                "description": row[4],
                "title": file_name,
                "language": "eng",
                "cost_center": cost_center,
                "date_range": row[5],
                "major_description": row[6],
                "minor_description": row[7],
                "reference_1": row[8]
            }
            
            file_metadata = {
                "publisher": publisher,
                "source_folder_path": file_path,
                "source_file_name": file_name,
                "dz_folder_path": folder_path,
                "dz_file_name": file_name,
                "file_tag": get_file_tag(file_name)
            }
            
            current_records.append({
                "operation": "create_record",
                "relation_id": relation_id,
                "record_metadata": record_metadata
            })
            
            current_records.append({
                "operation": "upload_new_file",
                "relation_id": relation_id,
                "file_metadata": file_metadata
            })
            
            if len(current_records) >= records_per_file * 2:
                output_file_path = os.path.join(
                    output_directory, f"{excel_base_name}_{sheet_name}_{file_count}.a360"
                )
                upload_count = write_records_to_file(output_file_path, current_records)
                total_upload_operations += upload_count
                current_records = []
                file_count += 1
                
    if current_records:
        output_file_path = os.path.join(
            output_directory, f"{excel_base_name}_{sheet_name}_{file_count}.a360"
        )
        upload_count = write_records_to_file(output_file_path, current_records)
        total_upload_operations += upload_count
        file_count += 1
        
    print(f"\n=== Processing Complete ===")
    print(f"Total manifest files created: {file_count}")
    print(f"Total upload_new_file operations: {total_upload_operations}")

def main():
    parser = argparse.ArgumentParser(description="Convert Excel file to JSON format.")
    parser.add_argument("input_file", type=str, help="Path to the Excel file")
    parser.add_argument("output_directory", type=str, help="Directory to save JSON files")
    
    try:
        if len(sys.argv) > 1:
            args = parser.parse_args([arg.strip('"') for arg in sys.argv[1:]])
        else:
            parser.print_help()
            sys.exit(1)
            
        print("\n=== Excel to JSON Converter ===")
        print(f"Input file: {args.input_file}")
        print(f"Output directory: {args.output_directory}")
        
        if not os.path.exists(args.input_file):
            print(f"Error: Input file '{args.input_file}' does not exist")
            sys.exit(1)
            
        os.makedirs(args.output_directory, exist_ok=True)
        
        process_excel_file(args.input_file, args.output_directory)
        
    except Exception as e:
        print(f"Error: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    main()
