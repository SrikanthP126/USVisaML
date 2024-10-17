import os
import json
import uuid
import openpyxl
import datetime
import time

# Security Classification Mapping
classification_map = {
    "Confidential": "C",
    "Highly Confidential": "HC",
    "Internal": "I",
    "Public": "P"
}

# Log file to track processed files
log_file_path = "/path/to/processed_files.log"

# Function to get the security classification abbreviation
def get_security_classification(value):
    return classification_map.get(value, "Unknown")

# Function to format the submission date
def get_submission_date():
    return datetime.datetime.utcnow().isoformat() + 'Z'

# Check if the file has already been processed
def is_file_processed(file_name):
    if not os.path.exists(log_file_path):
        return False
    with open(log_file_path, 'r') as log_file:
        processed_files = log_file.read().splitlines()
    return file_name in processed_files

# Log the processed file
def log_processed_file(file_name):
    with open(log_file_path, 'a') as log_file:
        log_file.write(f"{file_name}\n")

# Function to process Excel and convert to JSON
def process_excel_file(file_path, output_directory, records_per_file=100):
    if is_file_processed(file_path):
        print(f"File {file_path} has already been processed. Skipping.")
        return

    # Load Excel file with openpyxl
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        current_records = []
        file_counter = 1

        # Iterate over the rows in the sheet starting from row 9
        for row in sheet.iter_rows(min_row=9, values_only=True):
            if row[0] is None:
                continue  # Skip empty rows
            
            relation_id = str(uuid.uuid4())  # Create a unique relation ID

            record_metadata = {
                "record_class": row[2],  # Assuming column C9
                "publisher": sheet["B1"].value,  # B1
                "region": sheet["B4"].value,  # B4
                "recordDate": row[8],  # Date
                "provenance": sheet["D1"].value,  # D1
                "security_classification": get_security_classification(sheet["D4"].value),  # D4
                "contributor": sheet["D3"].value,  # D3
                "creator": sheet["B2"].value,  # B2
                "description": row[4],  # Description from row
                "language": "eng",
                "title": row[5],  # Title from row
                "Date Range": row[6],  # Date Range
                "Major Description": row[7],  # Major Description
                "Minor Description": row[8],  # Minor Description
                "Reference 1": row[9],  # Reference 1
                "submission_date": get_submission_date()
            }

            file_metadata = {
                "publisher": sheet["B1"].value,  # B1
                "source_folder_path": row[1],  # Source folder path
                "source_file_name": row[0],  # Source file name
                "dz_file_name": row[0],  # dz file name
                "file_tag": "zip"  # Assuming a default file tag
            }

            create_record = {
                "operation": "create_record",
                "relation_id": relation_id,
                "record_metadata": record_metadata
            }

            upload_new_file = {
                "operation": "upload_new_file",
                "relation_id": relation_id,
                "file_metadata": file_metadata
            }

            current_records.append(create_record)
            current_records.append(upload_new_file)

            if len(current_records) >= records_per_file * 2:
                output_file_path = os.path.join(output_directory, f'{sheet_name}_{file_counter}.a360')
                with open(output_file_path, 'w') as json_file:
                    for record in current_records:
                        json.dump(record, json_file, separators=(',', ':'))
                        json_file.write("\n")
                current_records = []
                file_counter += 1

        if current_records:
            output_file_path = os.path.join(output_directory, f'{sheet_name}_{file_counter}.a360')
            with open(output_file_path, 'w') as json_file:
                for record in current_records:
                    json.dump(record, json_file, separators=(',', ':'))
                    json_file.write("\n")
            current_records = []

    log_processed_file(file_path)
    print(f"Finished processing {file_path}")

# Function to check directory for new Excel files
def check_directory_for_new_files(excel_directory, output_directory, polling_interval=10):
    print(f"Watching directory: {excel_directory} for new Excel files...")
    while True:
        for file_name in os.listdir(excel_directory):
            file_path = os.path.join(excel_directory, file_name)
            if file_name.endswith(".xlsx") and not is_file_processed(file_name):
                print(f"Processing new file: {file_path}")
                process_excel_file(file_path, output_directory)
        time.sleep(polling_interval)

if __name__ == "__main__":
    excel_directory = "/path/to/excel_directory/"  # Update to your correct path
    output_directory = "/path/to/output_directory/"  # Update to your correct path
    check_directory_for_new_files(excel_directory, output_directory)
