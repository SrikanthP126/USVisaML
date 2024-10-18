import json
import os
import uuid
from openpyxl import load_workbook
from datetime import datetime

# File paths (you may change these to suit your environment)
excel_file_path = '/path/to/excel/file.xlsx'  # Update this with your Excel file path
output_directory = '/path/to/output/json/files/'  # Update this with your desired output directory

if not os.path.exists(output_directory):
    os.makedirs(output_directory)

# Helper function to format date to ISO 8601 format
def format_iso_date(excel_date):
    try:
        if isinstance(excel_date, str):
            if len(excel_date) == 7:  # Handling year-month format
                excel_date += '-01'
            return datetime.strptime(excel_date, "%Y-%m-%d").isoformat() + "-05:00"
    except (ValueError, TypeError):
        return None

# Helper function to get file tag based on file extension
def get_file_tag(file_name):
    ext = os.path.splitext(file_name)[1].lower()
    if ext == '.doc':
        return 'word'
    elif ext == '.pdf':
        return 'pdf'
    elif ext == '.zip':
        return 'zip'
    else:
        return 'unknown'

# Load the workbook
wb = load_workbook(excel_file_path, data_only=True)

# Helper function to write each record individually as one line in the file
def write_records_to_file(output_file_path, records):
    with open(output_file_path, 'w') as json_file:
        for record in records:
            json_file.write(json.dumps(record, separators=(',', ':')) + '\n')  # One record per line

# Process each sheet in the workbook
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    
    # Read metadata values from fixed cells (based on the Excel file layout)
    publisher = sheet['B1'].value
    region = sheet['B4'].value
    record_class = sheet['C7'].value
    provenance = sheet['D1'].value
    security_classification = sheet['D4'].value  # Needs abbreviation conversion
    security_classification_map = {"Confidential": "C", "Highly Confidential": "HC", "Internal": "I", "Public": "P"}
    security_classification = security_classification_map.get(security_classification, "I")  # Default to "I"

    creator = sheet['B2'].value
    contributor = sheet['D2'].value
    cost_center = sheet['B3'].value

    file_counter = 1
    record_counter = 1
    records_per_file = 100
    current_records = []

    # Iterate through rows of the sheet starting from row 9
    for row in sheet.iter_rows(min_row=9, values_only=True):
        if row[0] is None:
            break  # Stop when an empty row is encountered

        # Extract metadata from the row
        file_name = row[0]
        file_path = row[1]
        record_date = format_iso_date(row[2])
        description = row[3]
        date_range = row[4]
        major_description = row[5]
        minor_description = row[6]
        reference_1 = row[7]

        relation_id = str(uuid.uuid4())
        submission_date = datetime.now().isoformat() + "-05:00"  # System-generated submission date

        # Create the JSON records
        record_metadata = {
            "record_class": record_class,
            "submission_date": submission_date,
            "security_classification": security_classification,
            "contributor": contributor,
            "creator": creator,
            "description": description,
            "title": file_name,
            "language": "eng",
            "cost_center": cost_center,
            "date_range": date_range,
            "major_description": major_description,
            "minor_description": minor_description,
            "reference_1": reference_1
        }

        file_metadata = {
            "publisher": publisher,
            "source_folder_path": file_path,
            "source_file_name": file_name,
            "dz_file_name": file_name,
            "file_tag": get_file_tag(file_name)
        }

        # Append the 'create_record' and 'upload_new_file' operations
        current_records.append({
            "operation": "create_record",
            "relation_id": relation_id,
            "record_metadata": record_metadata
        })
        current_records.append({
            "operation": "upload_new_file",
            "relation_id": relation_id,
            "file_metadata": file_metadata
        })

        # Write to file if enough records are collected
        if len(current_records) >= records_per_file * 2:
            output_file_path = os.path.join(output_directory, f'{sheet_name}_{file_counter}.a360')
            write_records_to_file(output_file_path, current_records)
            current_records = []  # Reset for the next batch
            file_counter += 1

    # Write any remaining records to a final file
    if current_records:
        output_file_path = os.path.join(output_directory, f'{sheet_name}_{file_counter}.a360')
        write_records_to_file(output_file_path, current_records)

print("JSON files created successfully.")
