import os
import json
import uuid
import calendar
from openpyxl import load_workbook
from datetime import datetime

# Directory containing multiple Excel files
input_directory = '/path/to/excel/files/'
output_directory = '/path/to/json/output/'

if not os.path.exists(output_directory):
    os.makedirs(output_directory)

# Helper function to format the date to the last day of the month (if month/year is provided)
def format_iso_date(excel_date):
    try:
        if isinstance(excel_date, str):
            if len(excel_date) == 7:  # Year-Month format
                year, month = excel_date.split('-')
                # Get the last day of the month
                last_day = calendar.monthrange(int(year), int(month))[1]
                excel_date = f"{year}-{month}-{last_day}"
            return datetime.strptime(excel_date, "%Y-%m-%d").isoformat() + "-05:00"
    except (ValueError, TypeError):
        return None

# Helper function to get file tag based on file extension, including images and more types
def get_file_tag(file_name):
    ext = os.path.splitext(file_name)[1].lower()
    return {
        '.doc': 'word',
        '.pdf': 'pdf',
        '.zip': 'zip',
        '.jpg': 'image',
        '.jpeg': 'image',
        '.png': 'image',
        '.txt': 'text',
        '.csv': 'csv',
        # Add more types if needed
    }.get(ext, 'unknown')

# Helper function to write records to JSON file
def write_records_to_file(output_file_path, records):
    with open(output_file_path, 'w') as json_file:
        for record in records:
            json_file.write(json.dumps(record, separators=(',', ':')) + '\n')

# Function to process each Excel file in the directory
def process_excel_file(file_path):
    wb = load_workbook(file_path, data_only=True)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # Read metadata from fixed cells
        publisher = sheet['B1'].value
        region = sheet['B4'].value
        record_class = sheet['C7'].value
        provenance = sheet['D1'].value
        security_classification = sheet['D4'].value

        # Map security classification
        security_classification_map = {
            "Confidential": "C", 
            "Highly Confidential": "HC", 
            "Internal": "I", 
            "Public": "P"
        }
        security_classification = security_classification_map.get(security_classification, "I")

        # Other metadata fields
        creator = sheet['B2'].value
        contributor = sheet['D2'].value
        cost_center = sheet['B3'].value

        # Ensure the cost_center is always a 12-digit number
        cost_center = str(cost_center).zfill(12)

        file_counter = 1
        current_records = []
        records_per_file = 100

        # Iterate through the rows starting from row 9
        for row in sheet.iter_rows(min_row=9, values_only=True):
            if row[0] is None:
                break  # Stop when an empty row is encountered

            # Extract metadata from the row
            file_name = row[0]
            file_path = row[1]
            record_code = row[2]
            record_date = format_iso_date(row[5])

            relation_id = str(uuid.uuid4())
            submission_date = datetime.now().isoformat() + "-05:00"

            # Create the JSON records
            record_metadata = {
                "record_class": record_class,
                "publisher": publisher,
                "region": region,
                "recordDate": record_date,
                "provenance": provenance,
                "submission_date": submission_date,
                "security_classification": security_classification,
                "contributor": contributor,
                "creator": creator,
                "description": row[4],
                "title": file_name,
                "language": "eng",
                "cost_center": cost_center,
                "date_range": row[6],
                "major_description": row[7],
                "minor_description": row[8],
                "reference_1": row[9]
            }

            file_metadata = {
                "publisher": publisher,
                "source_folder_path": file_path,
                "source_file_name": file_name,
                "dz_file_name": file_name,
                "file_tag": get_file_tag(file_name)
            }

            # Append records to the list
            current_records.append({
                "operation": "create_record",
                "relation_id": relation_id,
                "record_metadata": record_metadata
            })
            current_records.append({
                "operation": "upload_new_file",
                "relation_id": relation_id,
                "file_metadata": file_metadata
            })

            # Write records to file if enough records are collected
            if len(current_records) >= records_per_file * 2:
                output_file_path = os.path.join(output_directory, f'{sheet_name}_{file_counter}.a360')
                write_records_to_file(output_file_path, current_records)
                current_records = []  # Reset for next batch
                file_counter += 1

        # Write any remaining records to a final file
        if current_records:
            output_file_path = os.path.join(output_directory, f'{sheet_name}_{file_counter}.a360')
            write_records_to_file(output_file_path, current_records)

# Function to process files in a directory
def process_directory(directory_path):
    excel_files = [f for f in os.listdir(directory_path) if f.endswith('.xlsx')]
    for excel_file in excel_files:
        file_path = os.path.join(directory_path, excel_file)
        process_excel_file(file_path)

    print("All Excel files have been processed. Script will now exit.")

# Main function to trigger the processing
if __name__ == "__main__":
    process_directory(input_directory)
