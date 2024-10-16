import os
import json
import uuid
from datetime import datetime
from openpyxl import load_workbook
import time

# Security Classification Mapping
classification_map = {
    "Confidential": "C",
    "Highly Confidential": "HC",
    "Internal": "I",
    "Public": "P"
}

# Function to get the security classification abbreviation
def get_security_classification(value):
    return classification_map.get(value, "Unknown")

# Function to format the submission date
def get_submission_date():
    return datetime.utcnow().isoformat() + 'Z'

# Log file to track processed files
log_file_path = "/path/to/processed_files.log"

# Check if the file has already been processed
def is_file_processed(file_name):
    if not os.path.exists(log_file_path):
        return False
    with open(log_file_path, 'r') as log_file:
        processed_files = log_file.read().splitlines()
    return file_name in processed_files

# Log the processed file
def log_processed_file(file_name):
    with open(log_file_path, 'a') as log_file:
        log_file.write(f"{file_name}\n")

# Process Excel file and convert to JSON
def process_excel_file(file_path, output_directory):
    wb = load_workbook(file_path)
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        current_records = []
        file_counter = 1

        # Skip first 8 rows (assuming metadata is in first 8 rows)
        for row in sheet.iter_rows(min_row=9, values_only=True):
            if row[0] is None:
                continue  # Skip rows with no data

            # Map Excel values to JSON fields
            record_class = row[2]
            record_date = row[8]
            description = row[4]
            date_range = row[5]
            major_description = row[6]
            minor_description = row[7]
            reference_1 = row[8]
            source_folder_path = row[1]
            source_file_name = row[0]
            dz_file_name = row[0]
            file_tag = "zip"  # Assuming zip for simplicity

            relation_id = str(uuid.uuid4())

            # Create JSON structure for create_record
            record_metadata = {
                "record_class": record_class,
                "publisher": sheet['B1'].value,
                "region": sheet['B4'].value,
                "recordDate": record_date,
                "provenance": sheet['D1'].value,
                "security_classification": get_security_classification(sheet['D4'].value),
                "contributor": sheet['D3'].value,
                "creator": sheet['B2'].value,
                "description": description,
                "language": "eng",
                "title": source_file_name,
                "Date Range": date_range,
                "Major Description": major_description,
                "Minor Description": minor_description,
                "Reference 1": reference_1,
                "submission_date": get_submission_date()
            }

            # Create JSON structure for upload_new_file
            file_metadata = {
                "publisher": sheet['B1'].value,
                "source_folder_path": source_folder_path,
                "source_file_name": source_file_name,
                "dz_file_name": dz_file_name,
                "file_tag": file_tag
            }

            current_records.append({
                "operation": "create_record",
                "relation_id": relation_id,
                "record_metadata": record_metadata
            })

            current_records.append({
                "operation": "upload_new_file",
                "relation_id": relation_id,
                "file_metadata": file_metadata
            })

        # Write records to JSON file
        if current_records:
            output_file_path = os.path.join(output_directory, f'{sheet_name}_{file_counter}.a360')
            with open(output_file_path, 'w') as json_file:
                for record in current_records:
                    json.dump(record, json_file, separators=(',', ':'))
                    json_file.write("\n")
            file_counter += 1

    # Log the processed file
    log_processed_file(file_path)
    print(f"Finished processing {file_path}")

# Check directory for new Excel files
def check_directory_for_new_files(excel_directory, output_directory):
    print(f"Watching directory: {excel_directory} for new Excel files...")

    processed = False
    for file_name in os.listdir(excel_directory):
        file_path = os.path.join(excel_directory, file_name)
        if file_name.endswith(".xlsx") and not is_file_processed(file_name):
            print(f"Processing new file: {file_path}")
            process_excel_file(file_path, output_directory)
            processed = True

    if not processed:
        print("No new files to process. Exiting...")
        return False  # Stop the loop if no new files
    return True  # Continue if files were processed

if __name__ == "__main__":
    excel_directory = "/ark/landing_zone/exceltojson/metadata_excel/"
    output_directory = "/ark/landing_zone/exceltojson/output/"
    
    while check_directory_for_new_files(excel_directory, output_directory):
        time.sleep(5
