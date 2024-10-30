import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# Define paths
input_folder = "idfexcel"  # Folder where input files are located
output_folder = "metadata_excel"  # Folder where output files will be saved

# Ensure the output folder exists
os.makedirs(output_folder, exist_ok=True)

# Helper functions for mappings
def extract_file_name(obj_name):
    return obj_name.split('/')[-1]

def extract_file_path(obj_name):
    return "/".join(obj_name.split('/')[:-1]) + "/"

def extract_last_date(date_range):
    return date_range.split('_')[-1]

# Load the input Excel file
input_file_path = os.path.join(input_folder, "input.xlsx")  # Replace "input.xlsx" with the actual file name
wb_input = load_workbook(input_file_path, data_only=True)

# Process each sheet
for sheet_name in wb_input.sheetnames:
    ws_input = wb_input[sheet_name]
    data = ws_input.values
    columns = next(data)
    df = pd.DataFrame(data, columns=columns)

    # Create a new workbook for output
    wb_output = Workbook()
    ws_output = wb_output.active
    ws_output.title = sheet_name

    # Add formatting and context up to row 8
    for row in ws_input.iter_rows(min_row=1, max_row=8, values_only=False):
        for cell in row:
            new_cell = ws_output.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell._style = cell._style  # Copy cell style

    # Apply field mappings and transformations
    df_output = pd.DataFrame({
        'file_name': df['Object Name'].apply(extract_file_name),
        'file_path': df['Object Name'].apply(extract_file_path),
        'record code': df['Record code'],
        'Last Date of Data': df['Date Range'].apply(extract_last_date),
        'Business Classification': df['Business Classification'],
        'Date Range': df['Date Range'],
        'Major Description': df['Object Name'],
        'Minor Description': df['App ID'],
        'Reference 1': df['Domain'],
        'Reference 2': df.get('Reference 2', ''),  # Leave blank if not present
        'Reference 3': df.get('Reference 3', ''),
        'Reference 4': df.get('Reference 4', ''),
        'Reference 5': df.get('Reference 5', ''),
        'Long Description': df.get('Long Description', ''),
    })

    # Write DataFrame to the output worksheet, starting from row 9
    for r_idx, row in enumerate(dataframe_to_rows(df_output, index=False, header=False), start=9):
        for c_idx, value in enumerate(row, start=1):
            ws_output.cell(row=r_idx, column=c_idx, value=value)

    # Save the output workbook
    output_file_path = os.path.join(output_folder, f"{sheet_name}_output.xlsx")
    wb_output.save(output_file_path)

print("Processing complete. Output files are saved in 'metadata_excel' folder.")
